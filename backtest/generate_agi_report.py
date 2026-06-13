"""
generate_agi_report.py — Gera Excel completo com resultados do AGI v12.
Gráficos, tabelas comparativas, análise por ativo/timeframe.
"""

import sys, csv, io, subprocess, os, json
from pathlib import Path
from datetime import datetime, time
import numpy as np
import pandas as pd

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, BarChart3D, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart.layout import Layout, ManualLayout

# ─── MT5 fetch ───────────────────────────────────────────────────────────────
WINE_PYTHON = os.path.expanduser("~/.wine/drive_c/Python311/python.exe")
FETCH_SCRIPT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mt5_fetch.py")

CONTRACT_SPECS = {
    "WIN$":  {"mult": 0.20, "name": "Mini Índice",    "margin": 155,  "slip_r": 1.0},
    "WDO$":  {"mult": 10.0, "name": "Mini Dólar",     "margin": 140,  "slip_r": 5.0},
    "BIT$":  {"mult": 0.50, "name": "Bitcoin",        "margin": 45,   "slip_r": 10.0},
    "DOL$":  {"mult": 50.0, "name": "Dólar Cheio",    "margin": 700,  "slip_r": 10.0},
    "WSP$":  {"mult": 2.5,  "name": "Micro S&P 500",  "margin": 100,  "slip_r": 2.5},
    "IND$":  {"mult": 1.00, "name": "Índice Bovespa", "margin": 775,  "slip_r": 5.0},
}

BEST_STRATEGY = {
    "WIN$": "BOLLINGER", "WDO$": "EMA_PULLBACK", "BIT$": "VWAP",
    "DOL$": "EMA_PULLBACK", "WSP$": "MACD_MOMENTUM", "IND$": "BOLLINGER",
}

COMMISSION = 1.2
CLOSE_HOUR, CLOSE_MINUTE = 16, 45
START_HOUR, START_MINUTE = 9, 5
ATR_PERIOD = 14

# ─── Colors ──────────────────────────────────────────────────────────────────
GREEN = "27AE60"
RED = "E74C3C"
BLUE = "2980B9"
ORANGE = "E67E22"
PURPLE = "8E44AD"
DARK_BG = "2C3E50"
LIGHT_BG = "ECF0F1"
HEADER_BG = "34495E"

# ─── Styles ──────────────────────────────────────────────────────────────────
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
green_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
red_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
title_font = Font(bold=True, size=14, color=DARK_BG)
subtitle_font = Font(bold=True, size=12, color=HEADER_BG)
bold_font = Font(bold=True, size=11)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

def style_data_cell(ws, row, col, value, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')
    if fmt:
        cell.number_format = fmt
    return cell


# ─── Indicators ──────────────────────────────────────────────────────────────
def calc_atr(df, period=14):
    h, l = df["high"], df["low"]
    c_prev = df["close"].shift(1)
    tr = pd.concat([h - l, (h - c_prev).abs(), (l - c_prev).abs()], axis=1).max(axis=1)
    return tr.rolling(period).mean()

def calc_vwap(df, period=20):
    typical = (df["high"] + df["low"] + df["close"]) / 3
    vol = df["tick_volume"].replace(0, 1)
    return (typical * vol).rolling(period).sum() / vol.rolling(period).sum()

def calc_ema(series, period):
    return series.ewm(span=period, adjust=False).mean()

def calc_rsi(series, period=14):
    delta = series.diff()
    gain = delta.where(delta > 0, 0.0)
    loss = -delta.where(delta < 0, 0.0)
    avg_gain = gain.ewm(alpha=1/period, min_periods=period).mean()
    avg_loss = loss.ewm(alpha=1/period, min_periods=period).mean()
    rs = avg_gain / avg_loss.replace(0, 1e-10)
    return 100 - (100 / (1 + rs))

def calc_adx(df, period=14):
    high, low, close = df["high"], df["low"], df["close"]
    plus_dm = high.diff()
    minus_dm = low.diff().mul(-1)
    plus_dm = plus_dm.where((plus_dm > minus_dm) & (plus_dm > 0), 0.0)
    minus_dm = minus_dm.where((minus_dm > plus_dm) & (minus_dm > 0), 0.0)
    tr = pd.concat([high - low, (high - close.shift(1)).abs(), (low - close.shift(1)).abs()], axis=1).max(axis=1)
    atr = tr.ewm(alpha=1/period, min_periods=period).mean()
    plus_di = 100 * plus_dm.ewm(alpha=1/period, min_periods=period).mean() / atr.replace(0, 1e-10)
    minus_di = 100 * minus_dm.ewm(alpha=1/period, min_periods=period).mean() / atr.replace(0, 1e-10)
    dx = 100 * (plus_di - minus_di).abs() / (plus_di + minus_di).replace(0, 1e-10)
    adx = dx.ewm(alpha=1/period, min_periods=period).mean()
    return adx, plus_di, minus_di

def calc_macd(df, fast=12, slow=26, signal=9):
    ema_fast = calc_ema(df["close"], fast)
    ema_slow = calc_ema(df["close"], slow)
    macd_line = ema_fast - ema_slow
    signal_line = calc_ema(macd_line, signal)
    histogram = macd_line - signal_line
    return macd_line, signal_line, histogram

def calc_bollinger(df, period=20, std=2.0):
    mid = df["close"].rolling(period).mean()
    std_val = df["close"].rolling(period).std()
    upper = mid + std * std_val
    lower = mid - std * std_val
    return upper, mid, lower


# ─── Strategy Checks ─────────────────────────────────────────────────────────
def check_vwap(price, cur_atr_pct, ema_f, ema_s, vwap_val, rsi_val):
    if vwap_val == 0: return None
    if cur_atr_pct < 0.0015: bm, sm = 1.0005, 0.9995
    elif cur_atr_pct < 0.003: bm, sm = 1.0015, 0.9985
    else: bm, sm = 1.002, 0.998
    d = None
    if price > vwap_val * bm: d = "BUY"
    elif price < vwap_val * sm: d = "SELL"
    if not d: return None
    if ema_f > 0 and ema_s > 0:
        if d == "BUY" and ema_f < ema_s: return None
        if d == "SELL" and ema_f > ema_s: return None
    if not pd.isna(rsi_val):
        if d == "BUY" and rsi_val > 85: return None
        if d == "SELL" and rsi_val < 15: return None
    return d

def check_ema_pullback(price, ema_f, ema_s, adx, pdi, mdi, rsi):
    if pd.isna(adx) or adx < 20: return None
    if pd.isna(ema_f) or pd.isna(ema_s) or ema_s == 0: return None
    up = ema_f > ema_s; dn = ema_f < ema_s
    if not up and not dn: return None
    if not pd.isna(pdi) and not pd.isna(mdi):
        if up and pdi < mdi: return None
        if dn and mdi < pdi: return None
    d = "BUY" if up else "SELL"
    if d == "BUY" and price < ema_s * 0.998: return None
    if d == "SELL" and price > ema_s * 1.002: return None
    if not pd.isna(rsi):
        if d == "BUY" and rsi > 80: return None
        if d == "SELL" and rsi < 20: return None
    return d

def check_macd_momentum(price, ema_f, ema_s, adx, rsi, h, ph, p2h):
    if pd.isna(adx) or adx < 15: return None
    if pd.isna(ema_f) or pd.isna(ema_s) or ema_s == 0: return None
    up = ema_f > ema_s; dn = ema_f < ema_s
    if not up and not dn: return None
    cu = ph <= 0 and h > 0; cd = ph >= 0 and h < 0
    mu = h > 0 and h > ph and ph > p2h; md = h < 0 and h < ph and ph < p2h
    d = None
    if up and (cu or mu):
        if not pd.isna(rsi) and rsi > 75: return None
        d = "BUY"
    elif dn and (cd or md):
        if not pd.isna(rsi) and rsi < 25: return None
        d = "SELL"
    if not d: return None
    if d == "BUY" and price < ema_s * 0.995: return None
    if d == "SELL" and price > ema_s * 1.005: return None
    return d

def check_bollinger(price, rsi, bu, bl):
    if bu == 0 or bl == 0 or pd.isna(bu): return None
    d = None
    if price <= bl: d = "BUY"
    elif price >= bu: d = "SELL"
    if not d: return None
    if not pd.isna(rsi):
        if d == "BUY" and rsi > 30: return None
        if d == "SELL" and rsi < 70: return None
    return d


# ─── Backtest Engine ─────────────────────────────────────────────────────────
def fetch(symbol, tf, n_bars):
    cmd = ["wine", WINE_PYTHON, FETCH_SCRIPT, "rates", symbol, tf, str(n_bars)]
    env = {**os.environ, "WINEDEBUG": "-all"}
    r = subprocess.run(cmd, capture_output=True, text=True, timeout=60, env=env)
    if r.returncode != 0 or not r.stdout.strip():
        return pd.DataFrame()
    reader = csv.reader(io.StringIO(r.stdout.strip()))
    headers = next(reader)
    rows = [x for x in reader if x]
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows, columns=headers)
    for c in ["open", "high", "low", "close", "tick_volume", "real_volume"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df["time"] = pd.to_datetime(df["time"].astype(int), unit="s")
    df = df.set_index("time")
    df["hour"] = df.index.hour
    df["minute"] = df.index.minute
    df["date"] = df.index.date
    return df[["open", "high", "low", "close", "tick_volume", "real_volume", "hour", "minute", "date"]].dropna(subset=["close"])


def backtest(df, symbol, tf, strategy, *, capital=1_000_000.0):
    spec = CONTRACT_SPECS[symbol]
    mult = spec["mult"]
    slip_r = spec["slip_r"]
    atr = calc_atr(df, ATR_PERIOD)
    _z = pd.Series(0.0, index=df.index)
    vwap = calc_vwap(df, 20) if strategy == "VWAP" else _z
    ema_f = calc_ema(df["close"], 9) if strategy in ("VWAP", "EMA_PULLBACK", "MACD_MOMENTUM") else _z
    ema_s = calc_ema(df["close"], 21) if strategy in ("VWAP", "EMA_PULLBACK", "MACD_MOMENTUM") else _z
    rsi = calc_rsi(df["close"], 14)
    adx_v, pdi, mdi = calc_adx(df, 14) if strategy in ("EMA_PULLBACK", "MACD_MOMENTUM") else (_z, _z, _z)
    _, _, hist = calc_macd(df) if strategy == "MACD_MOMENTUM" else (_z, _z, _z)
    bbu, bbm, bbl = calc_bollinger(df) if strategy == "BOLLINGER" else (_z, _z, _z)
    sl_mult = 1.0 if strategy == "BOLLINGER" else 1.5
    trail_act = 1.5; trail_dist = 0.5; cooldown = 300; max_daily = 8
    cash = capital; pos = 0; ep = 0.0; e_date = None; e_atr = 0.0
    best_p = 0.0; sl_p = 0.0; trail_on = False; sl_pts = 0; bars_in = 0
    trades = []; daily = {}; last_tt = None

    def _close(price, reason, date):
        nonlocal cash, pos, ep, e_date, best_p, sl_p, trail_on, e_atr, sl_pts, bars_in
        if pos == 0: return
        pnl = ((price - ep) * mult - slip_r - COMMISSION) if pos == 1 else ((ep - price) * mult - slip_r - COMMISSION)
        cash += pnl
        trades.append({"dir": "BUY" if pos == 1 else "SELL", "et": e_date, "xt": date,
                       "ep": ep, "xp": price, "pnl": pnl, "reason": reason, "bars": bars_in,
                       "symbol": symbol, "tf": tf, "strategy": strategy})
        pos = 0; ep = 0; best_p = 0; sl_p = 0; trail_on = False; bars_in = 0

    def _open(d, price, date, cur_atr):
        nonlocal cash, pos, ep, e_date, best_p, sl_p, trail_on, e_atr, sl_pts, last_tt
        if pos != 0: return False
        if last_tt is not None and (date - last_tt).total_seconds() < cooldown: return False
        dd = date.date() if hasattr(date, 'date') else date
        if daily.get(dd, 0) >= max_daily: return False
        raw = int(cur_atr * sl_mult); raw = max(raw, 50); raw = ((raw + 4) // 5) * 5
        if raw <= 0: return False
        pos = 1 if d == "BUY" else -1; ep = price; e_date = date; e_atr = cur_atr
        sl_pts = raw; best_p = price; trail_on = False
        sl_p = price - raw if pos == 1 else price + raw
        daily[dd] = daily.get(dd, 0) + 1; last_tt = date
        return True

    for i, (date, row) in enumerate(df.iterrows()):
        price = float(row["close"]); high = float(row["high"]); low = float(row["low"])
        hour = int(row["hour"]); minute = int(row["minute"])
        cur_atr = float(atr.iloc[i]) if i > 0 and not pd.isna(atr.iloc[i]) else 0
        if hour < START_HOUR or (hour == START_HOUR and minute < START_MINUTE): continue
        if pos != 0:
            bars_in += 1
            if pos == 1: best_p = max(best_p, high)
            else: best_p = min(best_p, low) if best_p > 0 else low
            profit = (best_p - ep) if pos == 1 else (ep - best_p)
            tfm = {"M5": 5, "M15": 15, "M30": 30, "H1": 60}.get(tf, 5)
            pm = bars_in * tfm
            if not trail_on and e_atr > 0 and profit >= trail_act * e_atr: trail_on = True
            if trail_on and e_atr > 0:
                td = trail_dist * e_atr
                if pos == 1:
                    ns = best_p - td
                    if ns > sl_p: sl_p = ns
                else:
                    ns = best_p + td
                    if ns < sl_p: sl_p = ns
            if sl_p > 0:
                if pos == 1 and low <= sl_p: _close(sl_p, "SL", date); continue
                elif pos == -1 and high >= sl_p: _close(sl_p, "SL", date); continue
            if hour > CLOSE_HOUR or (hour == CLOSE_HOUR and minute >= CLOSE_MINUTE):
                _close(price, "1645", date); continue
            continue
        if cur_atr <= 0: continue
        d = None
        if strategy == "VWAP":
            cv = float(vwap.iloc[i]) if not pd.isna(vwap.iloc[i]) else 0
            ef = float(ema_f.iloc[i]) if not pd.isna(ema_f.iloc[i]) else 0
            es = float(ema_s.iloc[i]) if not pd.isna(ema_s.iloc[i]) else 0
            cr = float(rsi.iloc[i]) if not pd.isna(rsi.iloc[i]) else 50
            cap = cur_atr / price if price > 0 else 0
            d = check_vwap(price, cap, ef, es, cv, cr)
        elif strategy == "EMA_PULLBACK":
            ef = float(ema_f.iloc[i]) if not pd.isna(ema_f.iloc[i]) else 0
            es = float(ema_s.iloc[i]) if not pd.isna(ema_s.iloc[i]) else 0
            ax = float(adx_v.iloc[i]) if not pd.isna(adx_v.iloc[i]) else 0
            pi = float(pdi.iloc[i]) if not pd.isna(pdi.iloc[i]) else 0
            mi = float(mdi.iloc[i]) if not pd.isna(mdi.iloc[i]) else 0
            cr = float(rsi.iloc[i]) if not pd.isna(rsi.iloc[i]) else 50
            d = check_ema_pullback(price, ef, es, ax, pi, mi, cr)
        elif strategy == "MACD_MOMENTUM":
            ef = float(ema_f.iloc[i]) if not pd.isna(ema_f.iloc[i]) else 0
            es = float(ema_s.iloc[i]) if not pd.isna(ema_s.iloc[i]) else 0
            ax = float(adx_v.iloc[i]) if not pd.isna(adx_v.iloc[i]) else 0
            cr = float(rsi.iloc[i]) if not pd.isna(rsi.iloc[i]) else 50
            ch = float(hist.iloc[i]) if not pd.isna(hist.iloc[i]) else 0
            ph = float(hist.iloc[i-1]) if i > 0 and not pd.isna(hist.iloc[i-1]) else 0
            p2h = float(hist.iloc[i-2]) if i > 1 and not pd.isna(hist.iloc[i-2]) else 0
            d = check_macd_momentum(price, ef, es, ax, cr, ch, ph, p2h)
        elif strategy == "BOLLINGER":
            bu = float(bbu.iloc[i]) if not pd.isna(bbu.iloc[i]) else 0
            bl = float(bbl.iloc[i]) if not pd.isna(bbl.iloc[i]) else 0
            cr = float(rsi.iloc[i]) if not pd.isna(rsi.iloc[i]) else 50
            d = check_bollinger(price, cr, bu, bl)
        if d: _open(d, price, date, cur_atr)
    if pos != 0: _close(float(df["close"].iloc[-1]), "FORCE", df.index[-1])
    return trades


def run():
    print("🔄 Buscando dados e rodando backtests...")
    
    timeframes = ["M5", "M15", "M30", "H1"]
    tf_bars = {"M5": 500, "M15": 500, "M30": 300, "H1": 200}
    
    all_results = []
    all_trades = []
    
    # Symbols to test (all 6 for comparison)
    symbols = ["WIN$", "WDO$", "BIT$", "DOL$", "WSP$", "IND$"]
    
    for symbol in symbols:
        strategy = BEST_STRATEGY[symbol]
        for tf in timeframes:
            n_bars = tf_bars[tf]
            df = fetch(symbol, tf, n_bars)
            if df.empty:
                print(f"  ❌ {symbol} {tf} — sem dados")
                continue
            
            trades = backtest(df, symbol, tf, strategy)
            n_days = df["date"].nunique()
            p0, p1 = float(df["close"].iloc[0]), float(df["close"].iloc[-1])
            price_change = (p1/p0 - 1) * 100
            
            if trades:
                n = len(trades)
                wins = sum(1 for t in trades if t["pnl"] > 0)
                losses = n - wins
                pnl = sum(t["pnl"] for t in trades)
                wr = wins / n * 100
                avg_win = np.mean([t["pnl"] for t in trades if t["pnl"] > 0]) if wins else 0
                avg_loss = np.mean([t["pnl"] for t in trades if t["pnl"] <= 0]) if losses else 0
                max_win = max([t["pnl"] for t in trades]) if trades else 0
                max_loss = min([t["pnl"] for t in trades]) if trades else 0
                gw = sum(t["pnl"] for t in trades if t["pnl"] > 0)
                gl = abs(sum(t["pnl"] for t in trades if t["pnl"] <= 0))
                pf = gw / gl if gl > 0 else 999
                pnl_per_day = pnl / n_days if n_days > 0 else pnl
                avg_bars = np.mean([t["bars"] for t in trades])
                
                # Exit reasons
                reasons = {}
                for t in trades:
                    r = t["reason"]
                    reasons[r] = reasons.get(r, 0) + 1
                
                # Win/loss streaks
                streak_w = 0; streak_l = 0; max_sw = 0; max_sl = 0
                for t in trades:
                    if t["pnl"] > 0:
                        streak_w += 1; streak_l = 0
                        max_sw = max(max_sw, streak_w)
                    else:
                        streak_l += 1; streak_w = 0
                        max_sl = max(max_sl, streak_l)
                
                # Cumulative PnL
                cum_pnl = []
                running = 0
                for t in trades:
                    running += t["pnl"]
                    cum_pnl.append(running)
                
                result = {
                    "symbol": symbol, "name": CONTRACT_SPECS[symbol]["name"],
                    "tf": tf, "strategy": strategy,
                    "n_trades": n, "n_wins": wins, "n_losses": losses,
                    "wr": wr, "pnl": pnl, "pf": pf,
                    "avg_win": avg_win, "avg_loss": avg_loss,
                    "max_win": max_win, "max_loss": max_loss,
                    "pnl_per_day": pnl_per_day, "n_days": n_days,
                    "avg_bars": avg_bars, "margin": CONTRACT_SPECS[symbol]["margin"],
                    "price_change": price_change,
                    "reasons": reasons, "max_streak_w": max_sw, "max_streak_l": max_sl,
                    "cum_pnl": cum_pnl,
                }
                all_results.append(result)
                all_trades.extend(trades)
                
                print(f"  ✅ {symbol} {tf}: {n}t, WR {wr:.1f}%, PnL R$ {pnl:+.1f}")
            else:
                print(f"  ⚪ {symbol} {tf}: 0 trades")
    
    # ─── Create Excel ───
    print("\n📊 Gerando Excel...")
    wb = Workbook()
    
    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 1: RESUMO EXECUTIVO
    # ═══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "📊 Resumo Executivo"
    ws1.sheet_properties.tabColor = BLUE
    
    # Title
    ws1.merge_cells("A1:N1")
    ws1["A1"] = "🧪 AGI v12 — Relatório de Performance"
    ws1["A1"].font = Font(bold=True, size=16, color=DARK_BG)
    ws1["A1"].alignment = Alignment(horizontal='center')
    
    ws1.merge_cells("A2:N2")
    ws1["A2"] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Backtest: M5/M15/M30/H1 | 6 Ativos"
    ws1["A2"].font = Font(size=10, color="666666")
    ws1["A2"].alignment = Alignment(horizontal='center')
    
    # Summary table
    headers = ["Ativo", "Nome", "Estratégia", "TF", "Trades", "Wins", "Losses", 
               "WR%", "PnL R$", "PF", "PnL/Dia", "Margin", "ROI%", "Price Δ%"]
    row = 4
    for col, h in enumerate(headers, 1):
        ws1.cell(row=row, column=col, value=h)
    style_header_row(ws1, row, len(headers))
    
    row = 5
    for r in all_results:
        roi = (r["pnl"] / r["margin"] * 100) if r["margin"] > 0 else 0
        vals = [r["symbol"], r["name"], r["strategy"], r["tf"],
                r["n_trades"], r["n_wins"], r["n_losses"],
                round(r["wr"], 1), round(r["pnl"], 1), round(r["pf"], 2),
                round(r["pnl_per_day"], 1), r["margin"],
                round(roi, 1), round(r["price_change"], 2)]
        for col, v in enumerate(vals, 1):
            cell = style_data_cell(ws1, row, col, v)
            if col == 9:  # PnL
                cell.fill = green_fill if v > 0 else red_fill
                cell.font = Font(bold=True, color=GREEN if v > 0 else RED)
            if col == 8:  # WR
                cell.fill = green_fill if v >= 50 else PatternFill(start_color="FEF9E7", fill_type="solid")
        row += 1
    
    # Totals
    row += 1
    ws1.cell(row=row, column=1, value="TOTAIS").font = bold_font
    total_pnl = sum(r["pnl"] for r in all_results)
    total_trades = sum(r["n_trades"] for r in all_results)
    total_margin = sum(r["margin"] for r in all_results)
    ws1.cell(row=row, column=5, value=total_trades).font = bold_font
    ws1.cell(row=row, column=9, value=round(total_pnl, 1)).font = Font(bold=True, size=12, color=GREEN if total_pnl > 0 else RED)
    ws1.cell(row=row, column=12, value=total_margin).font = bold_font
    ws1.cell(row=row, column=13, value=round(total_pnl/total_margin*100, 1) if total_margin > 0 else 0).font = bold_font
    
    # Column widths
    for col in range(1, len(headers) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 12
    
    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 2: GRÁFICO PnL POR ATIVO
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("📈 PnL por Ativo")
    ws2.sheet_properties.tabColor = GREEN
    
    # Data for chart
    ws2["A1"] = "Ativo"
    ws2["B1"] = "PnL Total R$"
    ws2["C1"] = "PnL/Dia R$"
    ws2["D1"] = "WR%"
    ws2["E1"] = "Profit Factor"
    style_header_row(ws2, 1, 5)
    
    # Aggregate by symbol
    sym_data = {}
    for r in all_results:
        s = r["symbol"]
        if s not in sym_data:
            sym_data[s] = {"pnl": 0, "pnl_day": 0, "trades": 0, "name": r["name"]}
        sym_data[s]["pnl"] += r["pnl"]
        sym_data[s]["pnl_day"] += r["pnl_per_day"]
        sym_data[s]["trades"] += r["n_trades"]
    
    row = 2
    for s, d in sorted(sym_data.items(), key=lambda x: x[1]["pnl"], reverse=True):
        ws2.cell(row=row, column=1, value=f"{s} ({d['name']})")
        ws2.cell(row=row, column=2, value=round(d["pnl"], 1))
        ws2.cell(row=row, column=3, value=round(d["pnl_day"], 1))
        row += 1
    
    # PnL Bar Chart
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "PnL Total por Ativo (R$)"
    chart1.y_axis.title = "PnL (R$)"
    chart1.x_axis.title = "Ativo"
    chart1.style = 10
    chart1.width = 20
    chart1.height = 12
    
    data = Reference(ws2, min_col=2, min_row=1, max_row=row-1)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=row-1)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.shape = 4
    
    ws2.add_chart(chart1, "A" + str(row + 2))
    
    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 3: HEATMAP POR TF
    # ═══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("🗓️ Heatmap TF")
    ws3.sheet_properties.tabColor = ORANGE
    
    ws3["A1"] = "Ativo"
    ws3["B1"] = "M5 PnL"
    ws3["C1"] = "M15 PnL"
    ws3["D1"] = "M30 PnL"
    ws3["E1"] = "H1 PnL"
    ws3["F1"] = "TOTAL"
    style_header_row(ws3, 1, 6)
    
    # Build heatmap data
    heatmap = {}
    for r in all_results:
        s = r["symbol"]
        if s not in heatmap:
            heatmap[s] = {}
        heatmap[s][r["tf"]] = r["pnl"]
    
    row = 2
    for s in ["BIT$", "DOL$", "IND$", "WIN$", "WSP$", "WDO$"]:
        if s not in heatmap:
            continue
        ws3.cell(row=row, column=1, value=f"{s} ({CONTRACT_SPECS[s]['name']})")
        total = 0
        for col_idx, tf in enumerate(["M5", "M15", "M30", "H1"], 2):
            val = heatmap[s].get(tf, 0)
            total += val
            cell = style_data_cell(ws3, row, col_idx, round(val, 1))
            if val > 0:
                cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            elif val < 0:
                cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
        total_cell = style_data_cell(ws3, row, 6, round(total, 1))
        total_cell.font = Font(bold=True, color=GREEN if total > 0 else RED)
        row += 1
    
    for col in range(1, 7):
        ws3.column_dimensions[get_column_letter(col)].width = 16
    
    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 4: DETALHES POR ATIVO
    # ═══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("📋 Detalhes")
    ws4.sheet_properties.tabColor = PURPLE
    
    headers4 = ["Ativo", "TF", "Estratégia", "Trades", "WR%", "PnL R$", "PF",
                "Avg Win", "Avg Loss", "Max Win", "Max Loss", "PnL/Dia",
                "SL", "1645", "Trail", "Force", "Max Streak W", "Max Streak L", "Days"]
    for col, h in enumerate(headers4, 1):
        ws4.cell(row=1, column=col, value=h)
    style_header_row(ws4, 1, len(headers4))
    
    row = 2
    for r in all_results:
        reasons = r.get("reasons", {})
        vals = [
            r["symbol"], r["tf"], r["strategy"],
            r["n_trades"], round(r["wr"], 1), round(r["pnl"], 1), round(r["pf"], 2),
            round(r["avg_win"], 1), round(r["avg_loss"], 1),
            round(r["max_win"], 1), round(r["max_loss"], 1),
            round(r["pnl_per_day"], 1),
            reasons.get("SL", 0), reasons.get("1645", 0),
            reasons.get("TRAIL", 0), reasons.get("FORCE", 0),
            r["max_streak_w"], r["max_streak_l"], r["n_days"]
        ]
        for col, v in enumerate(vals, 1):
            cell = style_data_cell(ws4, row, col, v)
            if col == 6:
                cell.fill = green_fill if v > 0 else red_fill
        row += 1
    
    for col in range(1, len(headers4) + 1):
        ws4.column_dimensions[get_column_letter(col)].width = 13
    
    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 5: TRADE-BY-TRADE
    # ═══════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("📜 Trades")
    ws5.sheet_properties.tabColor = DARK_BG
    
    headers5 = ["#", "Ativo", "TF", "Estratégia", "Dir", "Entry Time", "Exit Time",
                "Entry Price", "Exit Price", "PnL R$", "Reason", "Bars"]
    for col, h in enumerate(headers5, 1):
        ws5.cell(row=1, column=col, value=h)
    style_header_row(ws5, 1, len(headers5))
    
    row = 2
    for i, t in enumerate(all_trades, 1):
        et = t["et"].strftime("%d/%m %H:%M") if hasattr(t["et"], "strftime") else str(t["et"])
        xt = t["xt"].strftime("%d/%m %H:%M") if hasattr(t["xt"], "strftime") else str(t["xt"])
        vals = [i, t["symbol"], t["tf"], t["strategy"], t["dir"],
                et, xt, round(t["ep"], 2), round(t["xp"], 2),
                round(t["pnl"], 1), t["reason"], t["bars"]]
        for col, v in enumerate(vals, 1):
            cell = style_data_cell(ws5, row, col, v)
            if col == 10:
                cell.fill = green_fill if v > 0 else red_fill
        row += 1
    
    for col in range(1, len(headers5) + 1):
        ws5.column_dimensions[get_column_letter(col)].width = 14
    
    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 6: EQUITY CURVE (cumulative PnL per asset)
    # ═══════════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("💰 Equity Curve")
    ws6.sheet_properties.tabColor = GREEN
    
    # Build equity data per symbol
    sym_equity = {}
    for r in all_results:
        s = r["symbol"]
        if s not in sym_equity:
            sym_equity[s] = []
        cum = 0
        for p in r["cum_pnl"]:
            cum = p
            sym_equity[s].append(cum)
    
    # Write headers
    ws6["A1"] = "Trade #"
    col = 2
    sym_cols = {}
    for s in sorted(sym_equity.keys()):
        ws6.cell(row=1, column=col, value=s)
        sym_cols[s] = col
        col += 1
    style_header_row(ws6, 1, col - 1)
    
    # Find max trades
    max_trades = max(len(v) for v in sym_equity.values()) if sym_equity else 0
    
    for i in range(max_trades):
        ws6.cell(row=i+2, column=1, value=i+1)
        for s, vals in sym_equity.items():
            if i < len(vals):
                ws6.cell(row=i+2, column=sym_cols[s], value=round(vals[i], 1))
    
    # Equity Line Chart
    if max_trades > 0:
        chart2 = LineChart()
        chart2.title = "Equity Curve por Ativo"
        chart2.y_axis.title = "PnL Acumulado (R$)"
        chart2.x_axis.title = "Trade #"
        chart2.style = 10
        chart2.width = 25
        chart2.height = 14
        
        cats = Reference(ws6, min_col=1, min_row=2, max_row=max_trades+1)
        for s, col_idx in sym_cols.items():
            data = Reference(ws6, min_col=col_idx, min_row=1, max_row=max_trades+1)
            chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        
        ws6.add_chart(chart2, "A" + str(max_trades + 4))
    
    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 7: RECOMENDAÇÃO
    # ═══════════════════════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("🎯 Recomendação")
    ws7.sheet_properties.tabColor = GREEN
    
    ws7.merge_cells("A1:F1")
    ws7["A1"] = "🎯 RECOMENDAÇÃO AGI v12 — Ativos para Day Trade"
    ws7["A1"].font = Font(bold=True, size=14, color=DARK_BG)
    
    headers7 = ["Prioridade", "Ativo", "Estratégia", "PnL/Dia R$", "Confiança", "Ação"]
    for col, h in enumerate(headers7, 1):
        ws7.cell(row=3, column=col, value=h)
    style_header_row(ws7, 3, len(headers7))
    
    # Rank by total PnL
    ranked = sorted(sym_data.items(), key=lambda x: x[1]["pnl"], reverse=True)
    row = 4
    for priority, (s, d) in enumerate(ranked, 1):
        confidence = "🟢 Alta" if d["pnl"] > 0 and d["pnl_day"] > 100 else \
                     "🟡 Média" if d["pnl"] > 0 else "🔴 Baixa"
        action = "✅ OPERAR" if d["pnl"] > 0 else "❌ EVITAR"
        vals = [priority, f"{s} ({d['name']})", BEST_STRATEGY.get(s, "?"),
                round(d["pnl_day"], 1), confidence, action]
        for col, v in enumerate(vals, 1):
            cell = style_data_cell(ws7, row, col, v)
            if col == 6:
                cell.font = Font(bold=True, color=GREEN if "OPERAR" in str(v) else RED)
        row += 1
    
    for col in range(1, 7):
        ws7.column_dimensions[get_column_letter(col)].width = 18
    
    # Save
    output_path = "/home/bruno/Projects/Vibe-Trading/backtest/AGI_v12_Report.xlsx"
    wb.save(output_path)
    print(f"\n✅ Excel salvo: {output_path}")
    print(f"   {len(all_results)} combinações ativo/TF")
    print(f"   {len(all_trades)} trades no total")
    print(f"   7 abas: Resumo, Gráfico, Heatmap, Detalhes, Trades, Equity, Recomendação")


if __name__ == "__main__":
    run()
