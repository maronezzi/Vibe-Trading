#!/usr/bin/env python3
"""Generate Excel report from AGI results."""
import json
import os
import sys
from datetime import datetime

# Ensure openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

PROJECT = "/home/bruno/Projects/Vibe-Trading"

# Load data
with open(f"{PROJECT}/strategy_matrix_results.json") as f:
    matrix = json.load(f)

with open("/tmp/vt_agi_audit.json") as f:
    agi = json.load(f)

with open(f"{PROJECT}/vt_config.json") as f:
    config = json.load(f)

wb = Workbook()

# Colors
GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
BLUE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD = Font(bold=True, size=11)
TITLE = Font(bold=True, size=14)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def sc(ws, row, col, value, fmt=None, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')
    if fmt: cell.number_format = fmt
    if fill: cell.fill = fill
    return cell

# ═══ 1. Visao Geral AGI ═══
ws1 = wb.active
ws1.title = "Visao Geral AGI"
ws1.sheet_properties.tabColor = "4472C4"

ws1.merge_cells('A1:H1')
ws1.cell(row=1, column=1, value="AGI 17H - Resultado Completo").font = TITLE
ws1.cell(row=2, column=1, value=f"Data: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = BOLD
ws1.cell(row=3, column=1, value=f"Config: v{config.get('_version', '?')}").font = BOLD
ws1.cell(row=4, column=1, value=f"Periodo: {agi.get('period_days', 7)} dias").font = BOLD

row = 6
headers = ["Simbolo", "Trades", "WR %", "PnL R$", "Avg R$/Trade", "Status"]
for c, h in enumerate(headers, 1):
    ws1.cell(row=row, column=c, value=h)
style_header(ws1, row, len(headers))

perf = agi.get('performance', {})
by_sym = perf.get('by_symbol', {})
for sym in ["BIT", "WIN", "DOL", "IND", "WSP", "WDO"]:
    row += 1
    p = by_sym.get(sym, {})
    pnl = p.get('total_pnl', 0)
    wr = p.get('win_rate', 0)
    n = p.get('n_trades', 0)
    avg = p.get('avg_pnl', 0)
    status = "POSITIVO" if pnl > 0 else "NEGATIVO"
    fill = GREEN if pnl > 0 else RED
    
    sc(ws1, row, 1, sym, fill=fill)
    sc(ws1, row, 2, n)
    sc(ws1, row, 3, f"{wr:.1f}%", fill=fill)
    sc(ws1, row, 4, pnl, fmt='#,##0.00', fill=fill)
    sc(ws1, row, 5, avg, fmt='#,##0.00')
    sc(ws1, row, 6, status, fill=fill)

row += 1
total_pnl = sum(p.get('total_pnl', 0) for p in by_sym.values())
total_trades = sum(p.get('n_trades', 0) for p in by_sym.values())
sc(ws1, row, 1, "TOTAL", fill=BLUE)
sc(ws1, row, 2, total_trades, fill=BLUE)
sc(ws1, row, 3, "", fill=BLUE)
sc(ws1, row, 4, total_pnl, fmt='#,##0.00', fill=BLUE)
sc(ws1, row, 5, total_pnl/max(total_trades,1), fmt='#,##0.00', fill=BLUE)

for c in range(1, 7):
    ws1.column_dimensions[get_column_letter(c)].width = 16

# ═══ 2. Strategy Matrix ═══
ws2 = wb.create_sheet("Strategy Matrix")
ws2.sheet_properties.tabColor = "70AD47"

strategies = sorted(matrix.get('strategy_summary', {}).keys())
pairs = sorted(matrix.get('full_matrix', {}).keys())

ws2.cell(row=1, column=1, value="Matriz: Estrategia x Par (PnL R$)").font = TITLE

row = 3
ws2.cell(row=row, column=1, value="Par \\ Estrategia")
for c, strat in enumerate(strategies, 2):
    ws2.cell(row=row, column=c, value=strat)
style_header(ws2, row, len(strategies)+1)

full_matrix = matrix.get('full_matrix', {})
for r, pair in enumerate(pairs, row+1):
    ws2.cell(row=r, column=1, value=pair).font = BOLD
    ws2.cell(row=r, column=1).border = thin_border
    for c, strat in enumerate(strategies, 2):
        cell_data = full_matrix.get(pair, {}).get(strat, {})
        pnl = cell_data.get('pnl', 0)
        n = cell_data.get('n_trades', 0)
        fill = GREEN if pnl > 0 and n > 0 else RED if pnl < 0 and n > 0 else None
        sc(ws2, r, c, pnl, fmt='#,##0.00', fill=fill)

ws2.column_dimensions['A'].width = 14
for c in range(2, len(strategies)+2):
    ws2.column_dimensions[get_column_letter(c)].width = 16

# ═══ 3. Best per Pair ═══
ws3 = wb.create_sheet("Best per Pair")
ws3.sheet_properties.tabColor = "FFC000"

ws3.cell(row=1, column=1, value="Melhor Estrategia por Par").font = TITLE

row = 3
headers = ["Par", "Estrategia", "PnL R$", "Trades", "WR %", "Max DD R$"]
for c, h in enumerate(headers, 1):
    ws3.cell(row=row, column=c, value=h)
style_header(ws3, row, len(headers))

best = matrix.get('best_per_pair', {})
for pair in sorted(best.keys()):
    row += 1
    b = best[pair]
    fill = GREEN if b.get('pnl', 0) > 0 else RED
    sc(ws3, row, 1, pair, fill=fill)
    sc(ws3, row, 2, b.get('strategy', ''))
    sc(ws3, row, 3, b.get('pnl', 0), fmt='#,##0.00', fill=fill)
    sc(ws3, row, 4, b.get('n_trades', 0))
    sc(ws3, row, 5, f"{b.get('wr', 0):.0f}%")
    sc(ws3, row, 6, b.get('max_dd', 0), fmt='#,##0.00')

for c in range(1, 7):
    ws3.column_dimensions[get_column_letter(c)].width = 16

# ═══ 4. Strategy Ranking ═══
ws4 = wb.create_sheet("Strategy Ranking")
ws4.sheet_properties.tabColor = "ED7D31"

ws4.cell(row=1, column=1, value="Ranking de Estrategias").font = TITLE

row = 3
headers = ["#", "Estrategia", "PnL Total R$", "Wins", "Pairs", "Avg WR %"]
for c, h in enumerate(headers, 1):
    ws4.cell(row=row, column=c, value=h)
style_header(ws4, row, len(headers))

summary = matrix.get('strategy_summary', {})
ranked = sorted(summary.items(), key=lambda x: x[1]['total_pnl'], reverse=True)
for i, (strat, s) in enumerate(ranked, 1):
    row += 1
    fill = GREEN if s['total_pnl'] > 0 else RED if s['total_pnl'] < 0 else None
    medal = "1-Ouro" if i==1 else "2-Prata" if i==2 else "3-Bronze" if i==3 else str(i)
    sc(ws4, row, 1, medal)
    sc(ws4, row, 2, strat, fill=fill)
    sc(ws4, row, 3, s['total_pnl'], fmt='#,##0.00', fill=fill)
    sc(ws4, row, 4, s['n_wins'])
    sc(ws4, row, 5, s['n_pairs'])
    sc(ws4, row, 6, f"{s['avg_wr']:.1f}%")

for c in range(1, 7):
    ws4.column_dimensions[get_column_letter(c)].width = 18

# ═══ 5. Config Atual ═══
ws5 = wb.create_sheet("Config Atual")
ws5.sheet_properties.tabColor = "5B9BD5"

ws5.cell(row=1, column=1, value="Config vt_config.json - strategy_by_tf").font = TITLE
ws5.cell(row=2, column=1, value=f"Versao: v{config.get('_version', '?')}").font = BOLD

row = 4
headers = ["Par", "Estrategia", "Simbolo", "TF", "Status"]
for c, h in enumerate(headers, 1):
    ws5.cell(row=row, column=c, value=h)
style_header(ws5, row, len(headers))

sbt = config.get('strategy_by_tf', {})
disabled = set(config.get('disabled_timeframes', []))
for pair in sorted(sbt.keys()):
    row += 1
    parts = pair.split('_')
    sym = parts[0] if len(parts) > 0 else '?'
    tf = parts[1] if len(parts) > 1 else '?'
    is_disabled = pair in disabled
    fill = RED if is_disabled else GREEN
    status = "DESATIVADO" if is_disabled else "ATIVO"
    
    sc(ws5, row, 1, pair, fill=fill)
    sc(ws5, row, 2, sbt[pair])
    sc(ws5, row, 3, sym)
    sc(ws5, row, 4, tf)
    sc(ws5, row, 5, status, fill=fill)

# Params by TF
row += 2
ws5.cell(row=row, column=1, value="Params by TF").font = BOLD
row += 1
headers = ["Par", "sl_atr_mult", "cooldown_s", "max_daily_trades"]
for c, h in enumerate(headers, 1):
    ws5.cell(row=row, column=c, value=h)
style_header(ws5, row, len(headers))

pbtf = config.get('params_by_tf', {})
for pair in sorted(pbtf.keys()):
    row += 1
    p = pbtf[pair]
    sc(ws5, row, 1, pair)
    sc(ws5, row, 2, p.get('sl_atr_mult', '-'))
    sc(ws5, row, 3, p.get('cooldown_seconds', '-'))
    sc(ws5, row, 4, p.get('max_daily_trades', '-'))

for c in range(1, 6):
    ws5.column_dimensions[get_column_letter(c)].width = 18

# ═══ 6. Changes Applied ═══
ws6 = wb.create_sheet("Changes Applied")
ws6.sheet_properties.tabColor = "7030A0"

ws6.cell(row=1, column=1, value="Mudancas Aplicadas pelo AGI").font = TITLE

row = 3
headers = ["Simbolo", "Parametros", "Detalhe"]
for c, h in enumerate(headers, 1):
    ws6.cell(row=row, column=c, value=h)
style_header(ws6, row, len(headers))

changes = agi.get('changes_applied', [])
for ch in changes:
    row += 1
    sc(ws6, row, 1, ch.get('symbol', '?'))
    sc(ws6, row, 2, ', '.join(ch.get('params', [])))
    sc(ws6, row, 3, ch.get('detail', ''))

for c in range(1, 4):
    ws6.column_dimensions[get_column_letter(c)].width = 30

# ═══ 7. HALT Config ═══
ws7 = wb.create_sheet("HALT Config")
ws7.sheet_properties.tabColor = "FF0000"

ws7.cell(row=1, column=1, value="HALT Config por TF").font = TITLE

row = 3
headers = ["Par", "halt_duration_min", "max_consecutive_losses"]
for c, h in enumerate(headers, 1):
    ws7.cell(row=row, column=c, value=h)
style_header(ws7, row, len(headers))

hbtf = config.get('halt_by_tf', {})
for pair in sorted(hbtf.keys()):
    row += 1
    h = hbtf[pair]
    sc(ws7, row, 1, pair)
    sc(ws7, row, 2, h.get('halt_duration_minutes', '-'))
    sc(ws7, row, 3, h.get('max_consecutive_losses', '-'))

for c in range(1, 4):
    ws7.column_dimensions[get_column_letter(c)].width = 22

# Save
output = f"{PROJECT}/AGI_17h_resultado_completo.xlsx"
wb.save(output)
print(f"Excel gerado: {output}")
print(f"Abas: {len(wb.sheetnames)}")
for name in wb.sheetnames:
    print(f"  - {name}")
